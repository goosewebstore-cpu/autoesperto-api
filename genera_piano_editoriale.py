import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# FOGLIO 1: KEYWORD MAP
# ============================================================
ws1 = wb.active
ws1.title = "Keyword Map"

headers1 = ["Cluster", "Keyword Primaria", "Volume Stimato", "Difficoltà", "Search Intent", "Pagina Target", "Priority"]
ws1.append(headers1)

keyword_data = [
    # --- Cluster: Guide Budget ---
    ("Guide Budget", "migliori auto usate sotto i 10000 euro", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Guide Budget", "auto usate sotto i 5000 euro affidabili", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Guide Budget", "migliori auto usate sotto i 15000 euro", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Guide Budget", "auto familiare usata sotto i 20000 euro", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Guide Budget", "prima auto usata consigli per neopatentati", "Alto", "Media", "Informational", "Blog", "Alta"),
    
    # --- Cluster: Affidabilità e Problemi ---
    ("Affidabilità", "problemi comuni Audi A4 B8", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Affidabilità", "problemi Fiat Panda 169", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Affidabilità", "problemi Volkswagen Golf 6", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Affidabilità", "auto più affidabili usate classifica", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Affidabilità", "motori da evitare auto usate", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    
    # --- Cluster: Confronti ---
    ("Confronti", "Fiat Panda vs Volkswagen Up usate", "Medio", "Bassa", "Informational/Commerciale", "Blog + Tool Confronta", "Alta"),
    ("Confronti", "BMW Serie 3 vs Audi A4 usate quale scegliere", "Medio", "Bassa", "Informational/Commerciale", "Blog + Tool Confronta", "Alta"),
    ("Confronti", "SUV compatto usato confronto 2026", "Medio", "Media", "Informational", "Blog", "Media"),
    
    # --- Cluster: Elettriche e Ibride ---
    ("Elettriche/Ibride", "auto ibride usate migliori acquisto", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Elettriche/Ibride", "Tesla Model 3 usata problemi", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Elettriche/Ibride", "auto elettrica usata conviene", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Elettriche/Ibride", "svalutazione auto elettriche usate", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    
    # --- Cluster: Documenti e Burocrazia ---
    ("Documenti", "documenti necessari acquisto auto usata", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Documenti", "passaggio di proprietà auto costo", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Documenti", "come verificare auto non incidentata", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Documenti", "visura pra auto usata come funziona", "Medio", "Bassa", "Informational", "Blog", "Media"),
    
    # --- Cluster: Mercato e Svalutazione ---
    ("Mercato", "auto che si svalutano di meno classifica", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Mercato", "quanto si svaluta un auto in 3 anni", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Mercato", "miglior momento per comprare auto usata", "Medio", "Bassa", "Informational", "Blog", "Media"),
    
    # --- Cluster: Consigli Acquisto ---
    ("Consigli Acquisto", "come negoziare prezzo auto usata", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Consigli Acquisto", "cosa controllare prima di comprare auto usata", "Alto", "Media", "Informational", "Blog", "Alta"),
    ("Consigli Acquisto", "auto usata diesel o benzina 2026", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Consigli Acquisto", "chilometri auto usata quanti sono troppi", "Medio", "Bassa", "Informational", "Blog", "Alta"),
    ("Consigli Acquisto", "auto aziendale usata conviene", "Medio", "Bassa", "Informational", "Blog", "Media"),
    
    # --- Cluster: Per Marca (deep dives) ---
    ("Deep Dive Marca", "quale Fiat comprare usata guida completa", "Medio", "Bassa", "Informational", "Blog + Pagine Marca", "Media"),
    ("Deep Dive Marca", "quale BMW usata conviene Serie 1 3 5", "Medio", "Bassa", "Informational", "Blog + Pagine Modello", "Media"),
]

for row in keyword_data:
    ws1.append(row)

# Styling Foglio 1
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if row[0].row % 2 == 0:
        for cell in row:
            cell.fill = alt_fill

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 42
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 22
ws1.column_dimensions["F"].width = 22
ws1.column_dimensions["G"].width = 12

# ============================================================
# FOGLIO 2: PIANO EDITORIALE (20 Articoli)
# ============================================================
ws2 = wb.create_sheet("Piano Editoriale 20 Articoli")

headers2 = [
    "#", "Stato", "Cluster", "Titolo Articolo", "Keyword Primaria",
    "Keyword Secondarie", "Search Intent", "Lunghezza ( parole )",
    "Struttura Consigliata", "Link Interni (a)", "Link Interni (da)",
    "CTA", "Priority", "Data Pubblicazione", "Note"
]
ws2.append(headers2)

articoli = [
    (
        1, "Da scrivere", "Guide Budget",
        "Migliori Auto Usate sotto i 10.000€: La Guida Definitiva 2026",
        "migliori auto usate sotto i 10000 euro",
        "auto usate economiche affidabili, prima auto usata, auto usata sotto i 10000",
        "Informational/Commerciale", "2500",
        "Intro | Top 10 per fasce (3k-5k, 5k-7k, 7k-10k) | Tabella comparativa | Pro/contro per modello | Consigli finali",
        "Pagine valutazione modelli citati", "Homepage, pagine modello correlati",
        "Valuta la tua auto gratis — scopri il prezzo reale di mercato", "Alta", "Settimana 1",
        "Articolo pillar. Aggiornare ogni 6 mesi con nuovi prezzi."
    ),
    (
        2, "Da scrivere", "Guide Budget",
        "Prima Auto Usata: I 7 Modelli Perfetti per Neopatentati (2026)",
        "prima auto usata consigli per neopatentati",
        "auto neopatentati economiche, prima auto guida acquisto, auto usata per 18enni",
        "Informational", "2000",
        "Intro | Cosa cercare | Top 7 modelli | Assicurazione e costi | Checklist acquisto | Errore comuni",
        "Pagine modello, tool confronta", "Homepage, articoli budget",
        "Usa il nostro scanner AI per valutare l'auto che stai guardando", "Alta", "Settimana 2",
        "Target molto giovane. Usare tono colloquiale."
    ),
    (
        3, "Da scrivere", "Affidabilità",
        "Problemi Comuni Audi A4 B8 (2008-2015): Cosa Controllare Prima dell'Acquisto",
        "problemi comuni Audi A4 B8",
        "Audi A4 B8 difetti, affidabilità Audi A4, Audi A4 usata problemi",
        "Informational/Commerciale", "2200",
        "Intro | Storia modello | 7 problemi ricorrenti | Costi riparazione | Versioni consigliate | Versioni da evitare | Checklist prova",
        "Pagina /valutazione/audi/a4, tool confronta", "Articoli Audi, guide budget",
        "Scopri il prezzo reale di mercato di una Audi A4 B8", "Alta", "Settimana 3",
        "Template per altri modelli popolari. Inserire dati prezzi aggiornati."
    ),
    (
        4, "Da scrivere", "Affidabilità",
        "I 10 Motori da Evitare nell'Auto Usata (e Quelli che Durano per Sempre)",
        "motori da evitare auto usate",
        "motori problematici auto, motori affidabili usati, guida motori auto",
        "Informational", "2800",
        "Intro | Criteri valutazione | Top 5 motori da evitare (con motivi) | Top 5 motori affidabili | Tabella riassuntiva | FAQ",
        "Pagine modello con quei motori", "Homepage, articoli affidabilità",
        "Verifica il motore dell'auto che stai valutando — analisi AI gratuita", "Alta", "Settimana 4",
        "Contenuto evergreen. Può diventare aggiornamento annuale."
    ),
    (
        5, "Da scrivere", "Affidabilità",
        "Problemi Comuni Fiat Panda 169 (2003-2012): Guida all'Acquisto Usato",
        "problemi Fiat Panda 169",
        "Fiat Panda usata difetti, affidabilità Fiat Panda, Panda 169 problemi",
        "Informational/Commerciale", "2000",
        "Intro | Panoramica | 6 problemi principali | Costi manutenzione | Quale versione scegliere | Prezzo giusto",
        "Pagina /valutazione/fiat/panda", "Articoli Fiat, guide budget",
        "Scopri quanto vale la tua Panda usata — analisi gratuita", "Alta", "Settimana 5",
        "Modello molto cercato in Italia. Alto potenziale traffico."
    ),
    (
        6, "Da scrivere", "Confronti",
        "Fiat Panda vs Volkswagen Up: Quale City Car Usata Conviene?",
        "Fiat Panda vs Volkswagen Up usate",
        "confronto city car usate, Panda o Up, auto piccola usata",
        "Informational/Commerciale", "2200",
        "Intro | Criteri confronto | Scheda tecnica a confronto | Prezzi usato | Affidabilità | Spazi e consumi | Verdetto",
        "Pagine /valutazione/fiat/panda, /valutazione/volkswagen/up, /confronta", "Homepage, guide city car",
        "Metti a confronto i due modelli con un clic sul nostro tool", "Alta", "Settimana 6",
        "Replicare con altri confronti popolari (Yaris vs Fiesta, Golf vs A3)."
    ),
    (
        7, "Da scrivere", "Elettriche/Ibride",
        "Auto Ibride Usate: Le 8 Migliori per Qualità-Prezzo nel 2026",
        "auto ibride usate migliori acquisto",
        "ibrida usata conviene, Toyota Yaris hybrid usata, auto ibrida economica",
        "Informational/Commerciale", "2500",
        "Intro | Come funziona l'ibrido | Top 8 modelli | Confronto consumi reali | Costi batteria | Vantaggi vs diesel/benzina",
        "Pagine modello ibridi, tool confronta", "Homepage, guide elettriche",
        "Valuta il tuo ibrido usato — prezzo di mercato aggiornato", "Alta", "Settimana 7",
        "Trend in crescita. Aggiornare trimestralmente."
    ),
    (
        8, "Da scrivere", "Documenti",
        "Acquistare un'Auto Usata: I Documenti Essenziali e la Checklist Completa",
        "documenti necessari acquisto auto usata",
        "passaggio proprietà auto, contratto compravendita auto, visura pra",
        "Informational", "2200",
        "Intro | Documenti venditore | Documenti acquirente | Passaggio di proprietà | Visura PRA | Assicurazione | Checklist PDF scaricabile",
        "Homepage, pagine valutazione", "Tutti gli articoli acquisto",
        "Prima di comprare, scopri il valore reale dell'auto con AutoEsperto", "Alta", "Settimana 8",
        "Possibile upgrade: template PDF checklist scaricabile (lead magnet)."
    ),
    (
        9, "Da scrivere", "Mercato",
        "Le Auto che si Svalutano di Meno: Classifica 2026 (Dati Reali)",
        "auto che si svalutano di meno classifica",
        "svalutazione auto usate, auto che mantengono valore, investimento auto usata",
        "Informational", "2400",
        "Intro | Metodologia | Top 10 per fasce di prezzo | Grafico svalutazione | Perché certe auto resistono | FAQ",
        "Pagine valutazione modelli citati", "Homepage, tutti gli articoli",
        "Scopri quanto vale la tua auto oggi — analisi gratuita", "Alta", "Settimana 9",
        **"Contenuto da usare per digital PR: inviare a blog automotive."**
    ),
    (
        10, "Da scrivere", "Consigli Acquisto",
        "Come Negoziare il Prezzo di un'Auto Usata: 12 Strategie che Funzionano",
        "come negoziare prezzo auto usata",
        "trucchi comprare auto usata, ribassare prezzo auto, offerta auto usata",
        "Informational/Commerciale", "2200",
        "Intro | Preparazione prima del ritiro | 12 tecniche di negoziazione | Cosa NON dire | Quando accettare | Template messaggio",
        "Pagine valutazione (per avere il prezzo giusto in mano)", "Homepage, guide acquisto",
        "Conosci il prezzo di mercato prima di negoziare — prova AutoEsperto gratis", "Alta", "Settimana 10",
        "Molto condivisibile sui social. Possibile video correlato."
    ),
    (
        11, "Da scrivere", "Elettriche/Ibride",
        "Tesla Model 3 Usata: Problemi, Prezzi e Se Conviene nel 2026",
        "Tesla Model 3 usata problemi",
        "Tesla usata affidabilità, batteria Tesla usata, prezzo Tesla Model 3 usata",
        "Informational/Commerciale", "2300",
        "Intro | Panoramica mercato | Problemi noti | Costo batteria | Autonomia reale | Prezzi usato | Conviene?",
        "Pagina /valutazione/tesla/model-3", "Articoli elettriche, guide budget",
        "Scopri il valore reale di una Tesla Model 3 usata", "Alta", "Settimana 11",
        "Modello molto ricercato. Possibile alto traffico da ricerca."
    ),
    (
        12, "Da scrivere", "Consigli Acquosto",
        "Auto Usata: Quanti Chilometri Sono Troppi? La Risposta per Ogni Fas",
        "chilometri auto usata quanti sono troppi",
        "auto usata con tanti km, chilometraggio auto usata, usata con 200000 km",
        "Informational", "1800",
        "Intro | Regola generale | Per fasce di prezzo | Per tipo di motore | Come leggere il contachilometri | Quando i km non contano | FAQ",
        "Pagine valutazione per anno/km", "Guide acquisto, homepage",
        "Valuta l'auto in base a anno e km — analisi AI gratuita", "Alta", "Settimana 12",
        "Domanda molto comune. Ottimo per FAQ schema markup."
    ),
    (
        13, "Da scrivere", "Confronti",
        "BMW Serie 3 vs Audi A4 Usate: Quale Berlina Tedesca Scegliere?",
        "BMW Serie 3 vs Audi A4 usate quale scegliere",
        "confronto berline tedesche usate, Serie 3 o A4, BMW vs Audi usate",
        "Informational/Commerciale", "2200",
        "Intro | Storia confronto | Scheda tecnica | Affidabilità a confronto | Costi manutenzione | Prezzi usato | Verdetto per uso",
        "Pagine /valutazione/bmw/serie-3, /valutazione/audi/a4, /confronta", "Articoli premium, homepage",
        "Confronta prezzi e affidabilità nel nostro tool", "Alta", "Settimana 13",
        "Confronto classico. Alto potenziale di condivisione."
    ),
    (
        14, "Da scrivere", "Documenti",
        "Come Verificare se un'Auto è Incidentata: Guida Completa (PRA, Carfax, Prove)",
        "come verificare auto non incidentata",
        "auto incidentata come scoprire, visura pra incidenti, carfax italia",
        "Informational", "2000",
        "Intro | Segnali visibili | Documenti da chiedere | Visura PRA step-by-step | Servizi online | Test in officina | Checklist",
        "Homepage, pagine valutazione", "Tutti gli articoli acquisto",
        "Proteggi il tuo acquisto — valuta l'auto prima di comprare", "Alta", "Settimana 14",
        "Molto ricercato. Target ansioso/acquisto imminente."
    ),
    (
        15, "Da scrivere", "Consigli Acquisto",
        "Diesel o Benzina nel 2026: Cosa Scegliere per l'Auto Usata?",
        "auto usata diesel o benzina 2026",
        "diesel conviene 2026, benzina vs diesel usata, zone a traffico limitato diesel",
        "Informational", "2000",
        "Intro | Situazione normativa 2026 | Costi di esercizio | ZTL e diesel | Residuo valore | Per profilo di guida | Verdetto",
        "Pagine valutazione per tipo alimentazione", "Homepage, guide alimentazione",
        "Scopri il prezzo di mercato per diesel e benzina dello stesso modello", "Alta", "Settimana 15",
        "Evergreen. Aggiornare ogni anno con nuove normative."
    ),
    (
        16, "Da scrivere", "Mercato",
        "Quanto Si Svaluta un'Auto in 3 Anni? I Dati Reali per Marca (2026)",
        "quanto si svaluta un auto in 3 anni",
        "svalutazione auto per marca, calcolo svalutazione auto, auto che valgono di più",
        "Informational", "2000",
        "Intro | Come calcolare la svalutazione | Dati per segmento | Top 10 marche | Grafici | Come minimizzare | FAQ",
        "Pagine valutazione", "Tutti gli articoli, homepage",
        "Scopri quanto vale la tua auto oggi vs 3 anni fa", "Alta", "Settimana 16",
        **"Ottimo per digital PR. Dati unici = citazioni e link."**
    ),
    (
        17, "Da scrivere", "Guide Budget",
        "Migliori Auto Usate sotto i 5.000€: 10 Affidabili che Non Ti Lasciano a Piedi",
        "auto usate sotto i 5000 euro affidabili",
        "auto usata economica, auto 5000 euro usata, macchina usata poco prezzo",
        "Informational/Commerciale", "2200",
        "Intro | Criteri scelta | Top 10 modelli | Scheda per modello | Costi annuali | Dove cercare | Trappole da evitare",
        "Pagine valutazione modelli", "Homepage, guide budget",
        "Valuta ogni auto prima di comprare — prima analisi gratuita", "Alta", "Settimana 17",
        "Target price-sensitive. Possibile alto volume di ricerca."
    ),
    (
        18, "Da scrivere", "Elettriche/Ibride",
        "Auto Elettrica Usata: Conviene Davvero? Guida all'Acquisto 2026",
        "auto elettrica usata conviene",
        "elettrica usata problemi, batteria auto elettrica usata, ricarica auto elettrica",
        "Informational", "2400",
        "Intro | Stato mercato 2026 | Pro e contro usato elettrico | Costo batteria | Autonomia reale | Prezzi | Quando conviene",
        "Pagine modello elettrici", "Homepage, guide elettriche",
        "Valuta un'auto elettrica usata con il nostro scanner AI", "Alta", "Settimana 18",
        "Trend in forte crescita. Aggiornare semestralmente."
    ),
    (
        19, "Da scrivere", "Affidabilità",
        "Le Auto Più Affidabili Usate: La Nostra Classifica Basata su Dati Reali",
        "auto più affidabili usate classifica",
        "affidabilità auto usate, classifica auto usate, auto che non si rompono",
        "Informational", "2500",
        "Intro | Metodologia (dati annunci + segnalazioni) | Top 15 per segmento | Sorprese | Da evitare | Come verificare affidabilità",
        "Pagine valutazione modelli citati", "Homepage, tutti gli articoli",
        "Scopri l'affidabilità di qualsiasi auto — analisi gratuita", "Alta", "Settimana 19",
        **"Contenuto premium per digital PR. Dati esclusivi = link."**
    ),
    (
        20, "Da scrivere", "Deep Dive Marca",
        "Quale BMW Usata Conviene? Guida alla Scelta tra Serie 1, 3 e 5",
        "quale BMW usata conviene Serie 1 3 5",
        "BMW usata guida, Serie 3 usata problemi, BMW affidabile usata",
        "Informational/Commerciale", "2200",
        "Intro | Differenze tra Serie | Affidabilità per generazione | Costi manutenzione | Prezzi usato | Quale scegliere per uso | Verdetto",
        "Pagine /valutazione/bmw/*, tool confronta", "Articoli premium, homepage",
        "Confronta Serie 1, 3 e 5 — prezzi e affidabilità", "Alta", "Settimana 20",
        "Replicare con altre marche (Mercedes, Audi, Volkswagen)."
    ),
]

for row in articoli:
    ws2.append(row)

# Styling Foglio 2
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

priority_high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
priority_font_red = Font(color="9C0006")

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    # Colora righe priority Alta
    if row[12].value == "Alta":
        for cell in row:
            cell.fill = priority_high_fill
            cell.font = priority_font_red

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 48
ws2.column_dimensions["E"].width = 38
ws2.column_dimensions["F"].width = 38
ws2.column_dimensions["G"].width = 20
ws2.column_dimensions["H"].width = 14
ws2.column_dimensions["I"].width = 52
ws2.column_dimensions["J"].width = 28
ws2.column_dimensions["K"].width = 28
ws2.column_dimensions["L"].width = 42
ws2.column_dimensions["M"].width = 10
ws2.column_dimensions["N"].width = 14
ws2.column_dimensions["O"].width = 36

# Freeze top row
ws2.freeze_panes = "A2"

# ============================================================
# FOGLIO 3: CALENDARIO PUBBLICAZIONE (Roadmap)
# ============================================================
ws3 = wb.create_sheet("Calendario Roadmap")

headers3 = ["Settimana", "Data", "Articolo #", "Titolo", "Stato", "Azioni SEO Collegate", "Canale Promozione"]
ws3.append(headers3)

roadmap = [
    ("1", "Settimana 1", 1, "Migliori Auto Usate sotto i 10.000€", "Da scrivere", "Indexazione articolo, invio Search Console, link da homepage", "Blog, Newsletter, Social"),
    ("2", "Settimana 2", 2, "Prima Auto Usata: 7 Modelli per Neopatentati", "Da scrivere", "Link interno da articolo #1", "Blog, TikTok/Instagram"),
    ("3", "Settimana 3", 3, "Problemi Audi A4 B8", "Da scrivere", "Schema FAQ, link da pagina /valutazione/audi/a4", "Blog, Forum Audi"),
    ("4", "Settimana 4", 4, "Motori da Evitare nell'Auto Usata", "Da scrivere", "Contenuto evergreen, link da tutte le pagine modello", "Blog, Reddit/Forum"),
    ("5", "Settimana 5", 5, "Problemi Fiat Panda 169", "Da scrivere", "Schema FAQ, link da /valutazione/fiat/panda", "Blog, Gruppi Facebook Fiat"),
    ("6", "Settimana 6", 6, "Fiat Panda vs Volkswagen Up", "Da scrivere", "Link da /confronta, cross-linking", "Blog, Social"),
    ("7", "Settimana 7", 7, "Auto Ibride Usate: Top 8", "Da scrivere", "Link da pagine ibridi, aggiornamento trimestrale", "Blog, Newsletter"),
    ("8", "Settimana 8", 8, "Documenti Acquisto Auto Usata", "Da scrivere", "Lead magnet PDF, link da tutte le pagine", "Blog, Email, Social"),
    ("9", "Settimana 9", 9, "Auto che si Svalutano di Meno", "Da scrivere", "Digital PR: invio a blog automotive", "Blog, PR, Outreach"),
    ("10", "Settimana 10", 10, "Come Negoziare il Prezzo", "Da scrivere", "Video correlato, link da pagine prezzo", "Blog, YouTube, Social"),
    ("11", "Settimana 11", 11, "Tesla Model 3 Usata", "Da scrivere", "Link da /valutazione/tesla/model-3", "Blog, Forum Tesla"),
    ("12", "Settimana 12", 12, "Quanti Chilometri Sono Troppi?", "Da scrivere", "FAQ schema, link da pagine anno", "Blog, Social"),
    ("13", "Settimana 13", 13, "BMW Serie 3 vs Audi A4", "Da scrivere", "Link da /confronta, cross-marca", "Blog, Forum BMW/Audi"),
    ("14", "Settimana 14", 14, "Verificare Auto Incidentata", "Da scrivere", "Link da tutte le guide acquisto", "Blog, Email"),
    ("15", "Settimana 15", 15, "Diesel o Benzina 2026", "Da scrivere", "Evergreen, link da pagine per alimentazione", "Blog, Newsletter"),
    ("16", "Settimana 16", 16, "Svalutazione Auto in 3 Anni", "Da scrivere", "Digital PR, dati esclusivi", "Blog, PR, Outreach"),
    ("17", "Settimana 17", 17, "Auto sotto i 5.000€", "Da scrivere", "Link da homepage, target price-sensitive", "Blog, Social"),
    ("18", "Settimana 18", 18, "Auto Elettrica Usata Conviene?", "Da scrivere", "Link da pagine elettriche", "Blog, Newsletter"),
    ("19", "Settimana 19", 19, "Auto Più Affidabili Usate", "Da scrivere", "Digital PR, dati esclusivi", "Blog, PR, Outreach"),
    ("20", "Settimana 20", 20, "Quale BMW Usata Conviene?", "Da scrivere", "Link da pagine BMW, template per altre marche", "Blog, Forum BMW"),
]

for row in roadmap:
    ws3.append(row)

for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 42
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 42
ws3.column_dimensions["G"].width = 24

# ============================================================
# FOGLIO 4: BRIEF E STRATEGIA
# ============================================================
ws4 = wb.create_sheet("Strategia e Brief")

strategia_text = [
    ["STRATEGIA EDITORIALE AUTOESPERTO — BLOG 2026"],
    [],
    ["1. OBIETTIVO"],
    ["Portare il traffico organico da ~1.000 a 50.000+ visite/mese entro 12 mesi attraverso contenuti SEO"],
    ["che coprono l'intero funnel: informational (70%) → considerazione (20%) → conversione (10%)."],
    [],
    ["2. CLUSTER TEMATICI"],
    ["A. Guide Budget (20%) — Target: acquirenti in fase esplorativa, alto volume, bassa competizione"],
    ["B. Affidabilità e Problemi (25%) — Target: acquirenti in fase valutazione, intento commerciale forte"],
    ["C. Confronti (10%) — Target: decisori, alto tasso di conversione sul tool Confronta"],
    ["D. Elettriche e Ibride (15%) — Target: trend crescente, meno competizione, possibile posizionamento early"],
    ["E. Documenti e Burocrazia (10%) — Target: acquirenti prossimi all'acquisto, lead magnet opportunity"],
    ["F. Mercato e Svalutazione (10%) — Target: digital PR, link building, brand awareness"],
    ["G. Consigli Acquisto (10%) — Target: funnel medio-basso, alto engagement"],
    [],
    ["3. KPI MENSILI"],
    ["• Mese 1-2: 5 articoli pubblicati, 50 nuove parole chiave posizionate, 1.500 visite/mese"],
    ["• Mese 3-4: 10 articoli pubblicati, 200 nuove parole chiave posizionate, 5.000 visite/mese"],
    ["• Mese 5-6: 15 articoli pubblicati, 500 parole chiave posizionate, 15.000 visite/mese"],
    ["• Mese 7-12: 20+ articoli, 2.000 parole chiave, 50.000 visite/mese"],
    [],
    ["4. TEMPLATE ARTICOLO STANDARD"],
    ["• Intro (150 parole): hook con domanda/problema, promessa del contenuto"],
    ["• Sezioni H2 con keyword (ogni 300-400 parole)"],
    ["• Tabella comparativa o dati (ogni articolo deve avere almeno 1 tabella/elemento visivo)"],
    ["• FAQ accordion (minimo 4 domande, con schema FAQPage JSON-LD)"],
    ["• CTA in fondo: link a tool valutazione o confronto"],
    ["• Link interni: minimo 3 a pagine modello, minimo 2 da pagine modello all'articolo"],
    ["• Meta description: 150-160 caratteri con keyword + CTA"],
    [],
    ["5. PROMOZIONE"],
    ["• Settimana 1: pubblicazione + indexazione Search Console + link interni"],
    ["• Settimana 2: condivisione social (LinkedIn, Facebook gruppi auto, Reddit r/ItalyMotori)"],
    ["• Settimana 3: outreach a 5 blog/siti per citazione/link (solo articoli dati/esclusivi)"],
    ["• Settimana 4: aggiornamento se dati/prezzi cambiano"],
    [],
    ["6. AGGIORNAMENTO CONTENUTI"],
    ["• Articoli budget: ogni 6 mesi (prezzi cambiano)"],
    ["• Articoli elettriche/ibridi: ogni 3 mesi (mercato in evoluzione)"],
    ["• Articoli problemi modelli: ogni 12 mesi (se dati affidabilità stabili)"],
    ["• Articoli normative: ogni anno (ZTL, incentivi, ecc.)"],
]

for row in strategia_text:
    ws4.append(row)

ws4.column_dimensions["A"].width = 90

# Bold headers in strategy
bold_font = Font(bold=True, size=12)
for cell in ws4["A"]:
    if cell.value and cell.value.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "STRATEGIA")):
        cell.font = bold_font

# Save
output_path = r"C:\Users\noizz\OneDrive\Documents\TargaAI\autoesperto\Piano_Editoriale_AutoEsperto.xlsx"
wb.save(output_path)
print(f"File salvato: {output_path}")
